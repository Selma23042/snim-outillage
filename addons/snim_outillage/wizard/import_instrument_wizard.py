from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class InstrumentImportWizard(models.TransientModel):
    _name = 'snim.instrument.import.wizard'
    _description = 'Assistant import instruments Excel'

    file_data = fields.Binary(
        string='Fichier Excel',
        required=True,
        help='Selectionnez un fichier .xlsx contenant '
             'les instruments a importer.')
    file_name = fields.Char(string='Nom du fichier')
    import_mode = fields.Selection([
        ('create', 'Creer uniquement (ignorer les doublons)'),
        ('update', 'Mettre a jour si le code existe'),
    ], string='Mode d\'import', default='create', required=True)

    result_message = fields.Text(
        string='Resultat', readonly=True)

    def action_download_template(self):
        """Telecharger un template Excel vide."""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
        except ImportError:
            raise UserError(
                'openpyxl est requis. '
                'Installez : pip install openpyxl')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Instruments'

        header_font = Font(
            name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='2C3E50')
        center = Alignment(horizontal='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        headers = [
            ('code', 'Code *', 15),
            ('designation', 'Designation *', 30),
            ('capacite', 'Capacite *', 15),
            ('detenu_par', 'Detenu par *', 20),
            ('affectation', 'Affectation *', 20),
            ('periodicite', 'Periodicite (mois) *', 18),
            ('marque', 'Marque', 15),
            ('date_mise_en_service',
             'Date mise en service * (JJ/MM/AAAA)', 25),
            ('date_der_etalonnage',
             'Dernier etalonnage * (JJ/MM/AAAA)', 25),
            ('motif', 'Motif', 30),
        ]

        # Noms techniques en ligne 1 (cachee)
        for col_idx, (tech, label, width) in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=tech)

        # Labels en ligne 2
        for col_idx, (tech, label, width) in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
            letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[letter].width = width

        ws.row_dimensions[1].hidden = True
        ws.row_dimensions[2].height = 30

        # Exemple
        example = [
            'INST-001', 'Pied a coulisse', '0-150mm',
            'Atelier mecanique', 'Zone A', 12,
            'Mitutoyo', '01/01/2025', '15/03/2025',
            'Instrument neuf',
        ]
        for col_idx, val in enumerate(example, 1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font = Font(
                name='Arial', size=10, color='95A5A6',
                italic=True)

        # Instructions
        ws2 = wb.create_sheet('Instructions')
        instructions = [
            'INSTRUCTIONS D\'IMPORT',
            '',
            '1. Remplissez la feuille "Instruments"',
            '2. Les champs marques * sont obligatoires',
            '3. Les dates doivent etre au format '
            'JJ/MM/AAAA',
            '4. La periodicite est en mois (nombre entier)',
            '5. Ne modifiez pas la ligne 1 (cachee) '
            'ni la ligne 2 (en-tetes)',
            '6. L\'exemple en ligne 3 (gris) sera ignore',
            '',
            'Colonnes :',
            '  - Code : identifiant unique de l\'instrument',
            '  - Designation : nom complet',
            '  - Capacite : plage de mesure',
            '  - Detenu par : service ou personne',
            '  - Affectation : zone ou atelier',
            '  - Periodicite : frequence d\'etalonnage en mois',
            '  - Marque : fabricant (optionnel)',
            '  - Date mise en service : JJ/MM/AAAA',
            '  - Dernier etalonnage : JJ/MM/AAAA',
            '  - Motif : observations (optionnel)',
        ]
        for row_idx, text in enumerate(instructions, 1):
            cell = ws2.cell(row=row_idx, column=1, value=text)
            if row_idx == 1:
                cell.font = Font(
                    name='Arial', bold=True, size=14,
                    color='2C3E50')
            else:
                cell.font = Font(name='Arial', size=11)
        ws2.column_dimensions['A'].width = 60

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': 'Template_Import_Instruments_SNIM.xlsx',
            'type': 'binary',
            'datas': file_data,
            'mimetype': (
                'application/vnd.openxmlformats-'
                'officedocument.spreadsheetml.sheet'),
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}'
                   f'?download=true',
            'target': 'self',
        }

    def action_import(self):
        """Importer les instruments depuis le fichier Excel."""
        self.ensure_one()

        if not self.file_data:
            raise UserError('Veuillez selectionner un fichier.')

        if (self.file_name
                and not self.file_name.endswith(('.xlsx', '.xls'))):
            raise UserError(
                'Format non supporte. '
                'Utilisez un fichier .xlsx')

        try:
            import openpyxl
        except ImportError:
            raise UserError(
                'openpyxl est requis. '
                'Installez : pip install openpyxl')

        try:
            file_content = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(
                io.BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            raise UserError(
                f'Erreur a la lecture du fichier : {e}')

        created = 0
        updated = 0
        errors = []
        from datetime import datetime

        for row_idx, row in enumerate(
                ws.iter_rows(min_row=3, values_only=True), 3):
            if not row or not row[0]:
                continue

            # Ignorer l exemple (gris)
            code = str(row[0]).strip()
            if code == 'INST-001' and row_idx == 3:
                continue

            try:
                # Parser les dates
                def parse_date(val):
                    if not val:
                        return False
                    if isinstance(val, datetime):
                        return val.date()
                    if hasattr(val, 'date'):
                        return val
                    val_str = str(val).strip()
                    for fmt in (
                        '%d/%m/%Y', '%Y-%m-%d',
                        '%d-%m-%Y', '%d.%m.%Y',
                    ):
                        try:
                            return datetime.strptime(
                                val_str, fmt).date()
                        except ValueError:
                            continue
                    raise ValueError(
                        f'Format date invalide : {val}')

                vals = {
                    'code': code,
                    'designation': str(row[1] or '').strip(),
                    'capacite': str(row[2] or '').strip(),
                    'detenu_par': str(row[3] or '').strip(),
                    'affectation': str(row[4] or '').strip(),
                    'periodicite': int(row[5] or 12),
                    'marque': str(row[6] or '').strip() or False,
                    'date_mise_en_service': parse_date(row[7]),
                    'date_der_etalonnage': parse_date(row[8]),
                    'motif': str(row[9] or '').strip() or False,
                }

                # Validation
                required = [
                    'code', 'designation', 'capacite',
                    'detenu_par', 'affectation',
                    'date_mise_en_service',
                    'date_der_etalonnage',
                ]
                missing = [
                    f for f in required if not vals.get(f)]
                if missing:
                    errors.append(
                        f'Ligne {row_idx}: champs manquants '
                        f'({", ".join(missing)})')
                    continue

                existing = self.env['snim.instrument'].search(
                    [('code', '=', code)], limit=1)

                if existing:
                    if self.import_mode == 'update':
                        existing.write(vals)
                        updated += 1
                    else:
                        errors.append(
                            f'Ligne {row_idx}: code "{code}" '
                            f'existe deja (ignore)')
                else:
                    self.env['snim.instrument'].create(vals)
                    created += 1

            except Exception as e:
                errors.append(
                    f'Ligne {row_idx}: {str(e)}')

        # Message resultat
        msg_parts = []
        if created:
            msg_parts.append(
                f'{created} instrument(s) cree(s)')
        if updated:
            msg_parts.append(
                f'{updated} instrument(s) mis a jour')
        if errors:
            msg_parts.append(
                f'\n\nErreurs ({len(errors)}) :\n'
                + '\n'.join(errors[:20]))
            if len(errors) > 20:
                msg_parts.append(
                    f'... et {len(errors) - 20} autres erreurs')

        result = ' | '.join(msg_parts) if msg_parts else (
            'Aucune donnee importee.')

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import termine',
                'message': result,
                'type': 'success' if not errors else 'warning',
                'sticky': True,
                'next': {
                    'type': 'ir.actions.act_window',
                    'res_model': 'snim.instrument',
                    'view_mode': 'list,form',
                    'target': 'current',
                },
            },
        }
