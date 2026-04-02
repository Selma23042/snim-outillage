from odoo import models, fields, api
from odoo.exceptions import ValidationError, UserError
from dateutil.relativedelta import relativedelta
from datetime import date
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class Instrument(models.Model):
    _name = 'snim.instrument'
    _description = 'Instrument de mesure SNIM'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'code asc'
    _rec_name = 'display_name_computed'

    # ── Champs principaux ─────────────────────────────────
    code = fields.Char(
        string='Code', required=True, tracking=True,
        copy=False)
    designation = fields.Char(
        string='Designation', required=True, tracking=True)
    display_name_computed = fields.Char(
        string='Nom complet',
        compute='_compute_display_name_computed',
        store=True)
    capacite = fields.Char(
        string='Capacite', required=True)
    detenu_par = fields.Char(
        string='Detenu par', required=True)
    affectation = fields.Char(
        string='Affectation a', required=True)
    periodicite = fields.Integer(
        string='Periodicite (mois)',
        required=True, default=12)
    marque = fields.Char(string='Marque')
    date_mise_en_service = fields.Date(
        string='Date mise en service', required=True)
    date_der_etalonnage = fields.Date(
        string='Dernier etalonnage',
        required=True, tracking=True)
    date_pro_etalonnage = fields.Date(
        string='Prochain etalonnage',
        compute='_compute_date_pro',
        store=True, tracking=True)
    motif = fields.Text(string='Motif / Observations')

    statut = fields.Selection([
        ('en_attente', 'En attente'),
        ('conforme', 'Conforme'),
        ('non_conforme', 'Non conforme'),
        ('reforme', 'Reforme'),
    ], string='Statut', default='en_attente',
        tracking=True, required=True)

    etalonnage_status = fields.Selection([
        ('ok', 'A jour'),
        ('bientot', 'Bientot du'),
        ('depasse', 'Depasse'),
    ], string='Etat etalonnage',
        compute='_compute_etalonnage_status',
        store=True)

    jours_restants = fields.Integer(
        string='Jours restants',
        compute='_compute_etalonnage_status',
        store=True)

    # ── Relations ─────────────────────────────────────────
    fiche_vie_ids = fields.One2many(
        'snim.fiche.vie.materiel', 'instrument_id',
        string='Fiches de vie')
    fiche_vie_count = fields.Integer(
        compute='_compute_fiche_count',
        string='Nb fiches')
    reforme_ids = fields.One2many(
        'snim.reforme', 'instrument_id',
        string='Reformes')
    reforme_count = fields.Integer(
        compute='_compute_reforme_count',
        string='Nb reformes')

    # ── Champ pour import/export Excel ────────────────────
    color = fields.Integer(string='Couleur kanban')

    # ══════════════════════════════════════════════════════
    #   COMPUTED FIELDS
    # ══════════════════════════════════════════════════════

    @api.depends('code', 'designation')
    def _compute_display_name_computed(self):
        for rec in self:
            if rec.code and rec.designation:
                rec.display_name_computed = f"[{rec.code}] {rec.designation}"
            else:
                rec.display_name_computed = rec.designation or rec.code or 'Nouveau'

    @api.depends('date_der_etalonnage', 'periodicite')
    def _compute_date_pro(self):
        for rec in self:
            if rec.date_der_etalonnage and rec.periodicite:
                rec.date_pro_etalonnage = (
                    rec.date_der_etalonnage
                    + relativedelta(months=rec.periodicite)
                )
            else:
                rec.date_pro_etalonnage = False

    @api.depends('date_pro_etalonnage')
    def _compute_etalonnage_status(self):
        today = date.today()
        for rec in self:
            if not rec.date_pro_etalonnage:
                rec.etalonnage_status = 'ok'
                rec.jours_restants = 0
                continue
            delta = (rec.date_pro_etalonnage - today).days
            rec.jours_restants = delta
            if delta < 0:
                rec.etalonnage_status = 'depasse'
            elif delta <= 30:
                rec.etalonnage_status = 'bientot'
            else:
                rec.etalonnage_status = 'ok'

    def _compute_fiche_count(self):
        for rec in self:
            rec.fiche_vie_count = len(rec.fiche_vie_ids)

    def _compute_reforme_count(self):
        for rec in self:
            rec.reforme_count = len(rec.reforme_ids)

    # ══════════════════════════════════════════════════════
    #   CONTRAINTES
    # ══════════════════════════════════════════════════════

    _sql_constraints = [
        ('code_unique', 'UNIQUE(code)',
         'Le code instrument doit etre unique !'),
    ]

    @api.constrains('periodicite')
    def _check_periodicite(self):
        for rec in self:
            if rec.periodicite <= 0:
                raise ValidationError(
                    'La periodicite doit etre superieure a 0.')

    @api.constrains('date_mise_en_service', 'date_der_etalonnage')
    def _check_dates(self):
        for rec in self:
            if (rec.date_mise_en_service and rec.date_der_etalonnage
                    and rec.date_der_etalonnage < rec.date_mise_en_service):
                raise ValidationError(
                    'La date du dernier etalonnage ne peut pas '
                    'etre avant la date de mise en service.')

    # ══════════════════════════════════════════════════════
    #   ACTIONS BOUTONS
    # ══════════════════════════════════════════════════════

    def action_edit(self):
        """Ouvrir le formulaire en mode edition."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'snim.instrument',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
            'flags': {'mode': 'edit'},
        }

    def action_delete(self):
        """Supprimer l instrument avec confirmation."""
        self.ensure_one()
        code = self.code
        designation = self.designation
        self.unlink()
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Suppression reussie',
                'message': f'L\'instrument [{code}] {designation} '
                           f'a ete supprime.',
                'type': 'warning',
                'sticky': False,
                'next': {
                    'type': 'ir.actions.act_window',
                    'res_model': 'snim.instrument',
                    'view_mode': 'list,form',
                    'target': 'current',
                },
            },
        }

    def action_voir_fiches(self):
        """Voir les fiches de vie de cet instrument."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f'Fiches de vie - {self.code}',
            'res_model': 'snim.fiche.vie.materiel',
            'view_mode': 'list,form',
            'domain': [('instrument_id', '=', self.id)],
            'context': {'default_instrument_id': self.id},
        }

    def action_voir_reformes(self):
        """Voir les reformes de cet instrument."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f'Reformes - {self.code}',
            'res_model': 'snim.reforme',
            'view_mode': 'list,form',
            'domain': [('instrument_id', '=', self.id)],
            'context': {'default_instrument_id': self.id},
        }

    def action_set_conforme(self):
        """Marquer comme conforme."""
        self.write({'statut': 'conforme'})

    def action_set_non_conforme(self):
        """Marquer comme non conforme."""
        self.write({'statut': 'non_conforme'})

    def action_set_reforme(self):
        """Mettre en reforme et creer un enregistrement."""
        self.ensure_one()
        self.write({'statut': 'reforme'})
        self.env['snim.reforme'].create({
            'instrument_id': self.id,
            'date_reforme': date.today(),
            'motif': self.motif or 'Mise en reforme',
        })
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Reforme enregistree',
                'message': f'L\'instrument [{self.code}] a ete mis en reforme.',
                'type': 'success',
                'sticky': False,
            },
        }

    # ══════════════════════════════════════════════════════
    #   EXPORT EXCEL — Instruments a etalonner
    # ══════════════════════════════════════════════════════

    def action_export_etalonnage_excel(self):
        """
        Telecharger un fichier Excel contenant :
        - Les instruments dont l etalonnage est depasse
        - Les instruments dont l etalonnage est bientot du
        Avec la date de telechargement.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
        except ImportError:
            raise UserError(
                'Le module openpyxl est requis. '
                'Installez-le avec : pip install openpyxl')

        today = date.today()

        # Chercher tous les instruments a etalonner
        instruments = self.search([
            ('statut', '!=', 'reforme'),
            ('date_pro_etalonnage', '!=', False),
        ], order='date_pro_etalonnage asc')

        depasses = instruments.filtered(
            lambda r: r.date_pro_etalonnage < today)
        bientot = instruments.filtered(
            lambda r: today <= r.date_pro_etalonnage
            <= today + relativedelta(days=30))

        wb = openpyxl.Workbook()

        # ── Styles ────────────────────────────────────
        header_font = Font(
            name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill_red = PatternFill(
            'solid', fgColor='C0392B')
        header_fill_orange = PatternFill(
            'solid', fgColor='E67E22')
        header_fill_blue = PatternFill(
            'solid', fgColor='2C3E50')
        title_font = Font(
            name='Arial', bold=True, size=14, color='2C3E50')
        subtitle_font = Font(
            name='Arial', size=10, color='7F8C8D', italic=True)
        cell_font = Font(name='Arial', size=10)
        center = Alignment(
            horizontal='center', vertical='center',
            wrap_text=True)
        left = Alignment(
            horizontal='left', vertical='center',
            wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='BDC3C7'),
            right=Side(style='thin', color='BDC3C7'),
            top=Side(style='thin', color='BDC3C7'),
            bottom=Side(style='thin', color='BDC3C7'),
        )
        red_fill = PatternFill('solid', fgColor='FADBD8')
        orange_fill = PatternFill('solid', fgColor='FDEBD0')

        columns = [
            ('Code', 15),
            ('Designation', 30),
            ('Marque', 15),
            ('Capacite', 15),
            ('Detenu par', 20),
            ('Affectation', 20),
            ('Periodicite (mois)', 18),
            ('Dernier etalonnage', 18),
            ('Prochain etalonnage', 18),
            ('Jours de retard / restants', 25),
            ('Statut', 15),
        ]

        def write_sheet(ws, title, records, header_fill, row_fill):
            # Titre
            ws.merge_cells('A1:K1')
            c = ws['A1']
            c.value = f'SNIM — {title}'
            c.font = title_font
            c.alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:K2')
            c2 = ws['A2']
            c2.value = f'Date de telechargement : {today.strftime("%d/%m/%Y")}'
            c2.font = subtitle_font
            c2.alignment = Alignment(horizontal='center')

            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[4].height = 25

            # En-tetes
            for col_idx, (col_name, col_width) in enumerate(columns, 1):
                cell = ws.cell(row=4, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = col_width

            # Donnees
            for row_idx, inst in enumerate(records, 5):
                delta = (inst.date_pro_etalonnage - today).days
                if delta < 0:
                    jours_label = f'{abs(delta)} jours de retard'
                else:
                    jours_label = f'{delta} jours restants'

                data = [
                    inst.code,
                    inst.designation,
                    inst.marque or '',
                    inst.capacite or '',
                    inst.detenu_par or '',
                    inst.affectation or '',
                    inst.periodicite,
                    inst.date_der_etalonnage.strftime('%d/%m/%Y')
                    if inst.date_der_etalonnage else '',
                    inst.date_pro_etalonnage.strftime('%d/%m/%Y')
                    if inst.date_pro_etalonnage else '',
                    jours_label,
                    dict(inst._fields['statut'].selection).get(
                        inst.statut, ''),
                ]
                ws.row_dimensions[row_idx].height = 22
                for col_idx, val in enumerate(data, 1):
                    cell = ws.cell(
                        row=row_idx, column=col_idx, value=val)
                    cell.font = cell_font
                    cell.alignment = left if col_idx <= 6 else center
                    cell.border = thin_border
                    cell.fill = row_fill

            # Total
            total_row = len(records) + 5
            ws.merge_cells(f'A{total_row}:I{total_row}')
            c_total = ws.cell(
                row=total_row, column=1,
                value=f'Total : {len(records)} instrument(s)')
            c_total.font = Font(
                name='Arial', bold=True, size=11, color='2C3E50')
            c_total.alignment = center

        # ── Feuille 1 : Etalonnages depasses ──────────
        ws1 = wb.active
        ws1.title = 'Etalonnages depasses'
        write_sheet(
            ws1,
            'Instruments — Etalonnage DEPASSE',
            depasses,
            header_fill_red,
            red_fill,
        )

        # ── Feuille 2 : Bientot dus ───────────────────
        ws2 = wb.create_sheet('Bientot dus')
        write_sheet(
            ws2,
            'Instruments — Etalonnage dans les 30 prochains jours',
            bientot,
            header_fill_orange,
            orange_fill,
        )

        # ── Feuille 3 : Resume ────────────────────────
        ws3 = wb.create_sheet('Resume')
        ws3.merge_cells('A1:D1')
        ws3['A1'].value = 'SNIM — Resume etalonnages'
        ws3['A1'].font = title_font
        ws3['A1'].alignment = Alignment(horizontal='center')
        ws3.merge_cells('A2:D2')
        ws3['A2'].value = (
            f'Genere le {today.strftime("%d/%m/%Y a %H:%M")}')
        ws3['A2'].font = subtitle_font
        ws3['A2'].alignment = Alignment(horizontal='center')

        resume_data = [
            ('Indicateur', 'Nombre'),
            ('Total instruments (hors reformes)',
             len(instruments)),
            ('Etalonnages depasses', len(depasses)),
            ('Etalonnages dans les 30 jours', len(bientot)),
            ('Etalonnages a jour',
             len(instruments) - len(depasses) - len(bientot)),
        ]
        for row_idx, (label, val) in enumerate(resume_data, 4):
            ws3.cell(row=row_idx, column=1, value=label).font = (
                header_font if row_idx == 4
                else Font(name='Arial', size=11))
            ws3.cell(row=row_idx, column=1).fill = (
                header_fill_blue if row_idx == 4
                else PatternFill())
            ws3.cell(row=row_idx, column=2, value=val).font = (
                header_font if row_idx == 4
                else Font(name='Arial', bold=True, size=11))
            ws3.cell(row=row_idx, column=2).fill = (
                header_fill_blue if row_idx == 4
                else PatternFill())
            ws3.cell(row=row_idx, column=2).alignment = center
            ws3.cell(row=row_idx, column=1).border = thin_border
            ws3.cell(row=row_idx, column=2).border = thin_border

        ws3.column_dimensions['A'].width = 40
        ws3.column_dimensions['B'].width = 15

        # ── Generer le fichier ────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())
        output.close()

        filename = (
            f'SNIM_Etalonnages_{today.strftime("%Y%m%d")}.xlsx')

        # Creer l attachment
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'mimetype': (
                'application/vnd.openxmlformats-'
                'officedocument.spreadsheetml.sheet'),
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}'
                   f'?download=true',
            'target': 'self',
        }

    # ══════════════════════════════════════════════════════
    #   IMPORT EXCEL
    # ══════════════════════════════════════════════════════

    def action_import_excel(self):
        """Ouvrir le wizard d import Excel."""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Importer des instruments (Excel)',
            'res_model': 'snim.instrument.import.wizard',
            'view_mode': 'form',
            'target': 'new',
        }

    # ══════════════════════════════════════════════════════
    #   CRON : Alerte etalonnage
    # ══════════════════════════════════════════════════════

    @api.model
    def _cron_check_etalonnage(self):
        """Verifie les etalonnages et cree des activites."""
        today = date.today()
        soon = today + relativedelta(days=30)
        instruments = self.search([
            ('statut', '!=', 'reforme'),
            ('date_pro_etalonnage', '<=', soon),
            ('date_pro_etalonnage', '!=', False),
        ])
        for inst in instruments:
            inst.activity_schedule(
                'mail.mail_activity_data_todo',
                summary=f'Etalonnage a planifier : {inst.code}',
                note=(
                    f'L\'instrument {inst.code} - '
                    f'{inst.designation} necessite un '
                    f'etalonnage avant le '
                    f'{inst.date_pro_etalonnage}.'),
            )